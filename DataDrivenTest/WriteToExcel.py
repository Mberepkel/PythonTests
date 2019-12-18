import openpyxl
path = "C:/Users/VICOMA/Desktop/Work/IT Training/DDTest Sheet 3.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet=workbook.get_sheet_by_name("Sheet1")
sheet = workbook.active
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value="welcome"
        ##sheet.cell(row=r,column=c,value="welcome") another way you can write it

workbook.save(path)