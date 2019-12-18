import openpyxl

## One way to read excel file is to use Pandas with following commands
## import pandas as pd # another method to read from excel
## df = pd.read_excel ("C:/Users/VICOMA/Desktop/Work/IT Training/DDTEST SHEET.xlsx")
## print(df.head()) , () represents how many rows to display

path = ("C:/Users/VICOMA/Desktop/Work/IT Training/DDTEST SHEET.xlsx")
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
#sheet=workbook.get_sheet_by_name("Sheet2") # to open another sheet on the workbook

rows= sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print (sheet.cell(row=r,column=c).value, end="                ")


    print()




